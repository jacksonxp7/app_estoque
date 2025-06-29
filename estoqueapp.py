import datetime
import openpyxl
import produtos
import webbrowser as wb
import os
import sqlite3
from tkinter import ttk
from tkinter import *
from tkinter.filedialog import askopenfilename
from PIL import ImageTk, Image
import json

verde = '#009688'
tela1 = Tk()
tela1.resizable(False, False)
###################################################################################################################
#                                                           data
dataa = datetime.date.today()
data_verse = f'{dataa.day}/{dataa.month}/{dataa.year}'
hourrr = datetime.datetime.now()
hora = f'{hourrr.hour:02d}:{hourrr.minute:02d}:{hourrr.second:02d}'
hoje = f'{dataa.day}/0{dataa.month}/{dataa.year}\n {hora}'

###################################################################################################################

#                                                        imagens botões
pagoutodo=PhotoImage(file='front_end/botoes_menu/pagouvalor.png')
salvar = PhotoImage(file='front_end/botoes_menu/salvar.png')
salvarparcialmente = PhotoImage(file='front_end/botoes_menu/salvarparcialmente.png')
deletar  = PhotoImage(file='front_end/botoes_menu/deletar.png')
deletarrx = PhotoImage(file='front_end/botoes_menu/deletarr.png')
informaçoes = PhotoImage(file='front_end/botoes_menu/informaçoes.png')
anotaçoes = PhotoImage(file='front_end/botoes_menu/anotaçoes.png')
valor_pago = PhotoImage(file='front_end/botoes_menu/valor pago.png')
valor_pago1 = PhotoImage(file='front_end/botoes_menu/valor pago1.png')
pesquisar = PhotoImage(file='front_end/botoes_menu/pesquisar.png')
limpar_pesquisa = PhotoImage(file='front_end/botoes_menu/limpar pesquisa.png')
atutu = PhotoImage(file='front_end/labelinfo/atualizar.png')
atututu = PhotoImage(file='front_end/labelinfo/atualizar_estoque.png')
#                                                           imagens botões redes
facebook = PhotoImage(file='front_end/botoes_redes/facebook.png')
whatsapp = PhotoImage(file='front_end/botoes_redes/Whatsapp.png')
instagram = PhotoImage(file='front_end/botoes_redes/instagram.png')
moeda = PhotoImage(file='front_end/botoes_redes/moeda.png')
historico = PhotoImage(file='front_end/botoes_redes/historico.png')
atualizar_preços = PhotoImage(file='front_end/botoes_redes/atualizar_preços.png')
perfume = PhotoImage(file='front_end/botoes_redes/perfume.png')
e = PhotoImage(file='front_end/botoes_redes/e.png')
bd = PhotoImage(file='front_end/botoes_redes/bd.png')
#
produto = PhotoImage(file='front_end/labelinfo/produtos.png')
preço = PhotoImage(file='front_end/labelinfo/preços.png')
###################################################################################################################


def load():
    with open('elementos\setingss.json', 'r') as pi:
        resf = json.load(pi)

    try:
        # 1
        foto_procurada1 = Image.open(resf[0]['fotos1'])
        foto_procurada1 = foto_procurada1.resize((230, 230))
        photo1 = ImageTk.PhotoImage(foto_procurada1)
        foto1.configure(image=photo1)
        foto1.image = photo1
    except:
        print('nao encontrada')
        foto_procurada1 = Image.open('front_end/h.jpg')
        foto_procurada1 = foto_procurada1.resize((230, 230))
        photo1 = ImageTk.PhotoImage(foto_procurada1)
        foto1.configure(image=photo1)
        foto1.image = photo1

    try:
        # 2
        foto_procurada2 = Image.open(resf[0]['fotos2'])
        foto_procurada2 = foto_procurada2.resize((230, 230))
        photo2 = ImageTk.PhotoImage(foto_procurada2)
        foto2.configure(image=photo2)
        foto2.image = photo2
    except:
        print('nao encontrada')
        foto_procurada2 = Image.open('front_end/h.jpg')
        foto_procurada2 = foto_procurada2.resize((230, 230))
        photo2 = ImageTk.PhotoImage(foto_procurada2)
        foto2.configure(image=photo2)
        foto2.image = photo2

    try:
        # 3
        foto_procurada3 = Image.open(resf[0]['fotos3'])
        foto_procurada3 = foto_procurada3.resize((230, 230))
        photo3 = ImageTk.PhotoImage(foto_procurada3)
        foto3.configure(image=photo3)
        foto3.image = photo3
    except:
        print('nao encontrada')
        foto_procurada3 = Image.open('front_end/h.jpg')
        foto_procurada3 = foto_procurada3.resize((230, 230))
        photo3 = ImageTk.PhotoImage(foto_procurada3)
        foto3.configure(image=photo3)
        foto3.image = photo3

    try:
        # direita
        foto_procurada4 = Image.open(resf[0]['fotos4'])
        foto_procurada4 = foto_procurada4.resize((230, 470))
        photo4 = ImageTk.PhotoImage(foto_procurada4)
        foto4.configure(image=photo4)
        foto4.image = photo4
    except:
        print('nao encontrada')
        foto_procurada4 = Image.open('front_end/h.jpg')
        foto_procurada4 = foto_procurada4.resize((230, 230))
        photo4 = ImageTk.PhotoImage(foto_procurada4)
        foto4.configure(image=photo4)
        foto4.image = photo4


def mudar_tema():
    with open('elementos\setingss.json', 'r') as pi:
        resf = json.load(pi)

    def sav1():
        a = askopenfilename()
        resf[0]['fotos1'] = a
        with open('elementos/setingss.json', 'w') as arquivo_json:
            json.dump(resf, arquivo_json)
        load()
        ab.destroy()

    def sav2():
        a = askopenfilename()
        resf[0]['fotos2'] = a
        with open('elementos/setingss.json', 'w') as arquivo_json:
            json.dump(resf, arquivo_json)
        load()
        ab.destroy()

    def sav3():
        a = askopenfilename()
        resf[0]['fotos3'] = a
        with open('elementos/setingss.json', 'w') as arquivo_json:
            json.dump(resf, arquivo_json)
        load()
        ab.destroy()

    def sav4():
        a = askopenfilename()
        resf[0]['fotos4'] = a
        with open('elementos/setingss.json', 'w') as arquivo_json:
            json.dump(resf, arquivo_json)
        load()
        ab.destroy()

    ab = Toplevel(tela1)
    ab.resizable(False, False)
    ab.iconbitmap('front_end/botoes_redes/E.ico')
    ab.title('pago parcialmente')

    ab.configure(bg='#BBE0E5')
    # ab.geometry('600x480+650+300')
    btsf1 = Button(ab, text='foto 1', command=sav1, bg='black',
                   fg='white', font=('arial', '29'), width=30)
    btsf1.pack()
    btsf2 = Button(ab, text='foto 2', command=sav2, bg='black',
                   fg='white', font=('arial', '29'),width=30)
    btsf2.pack()
    btsf3 = Button(ab, text='foto 3', command=sav3, bg='black',
                   fg='white', font=('arial', '29'),width=30)
    btsf3.pack()
    btsf4 = Button(ab, text='foto direita', command=sav4,
                   bg='black', fg='white', font=('arial', '29'),width=30)
    btsf4.pack()


############################


def conex():  # dar os valores ao combobox
    global res
    res = []
    conn = sqlite3.connect('elementos/doces.db')
    c = conn.cursor()
    c.execute(f"SELECT produtos FROM produtos")
    resultados = c.fetchall()
    for xx in resultados:
        res.append(xx[0])

    # print(res)
    conn.commit()
    conn.close()
    global ent2s
    ent2s = ttk.Combobox(tela, font=('(arial', 29), values=res)

    ent2s.configure(width=11)
    ent2s.place(x=549, y=92)


def corrigir_resultado(varr):
    varr = str(varr)
    if ',' in varr or '.' in varr:
        filtro1 = varr.replace(',', '').replace('.', '')
        tamanho = len(str(filtro1))
        print(tamanho)

        valor = str(filtro1)
        # print("valor", valor)
        if tamanho == 9:
            v1 = f'{valor[0]}.{valor[1]}{valor[2]}{valor[3]}.{valor[4]}{valor[5]}{valor[6]},{valor[7]}{valor[8]}'
            print(v1)
            return v1
        elif tamanho == 8:
            v1 = f'{valor[0]}{valor[1]}{valor[2]}.{valor[3]}{valor[4]}{valor[5]},{valor[6]}{valor[7]}'
            print(v1)
            return v1
        elif tamanho == 7:
            v1 = f'{valor[0]}{valor[1]}.{valor[2]}{valor[3]}{valor[4]},{valor[5]}{valor[6]}'
            print(v1)
            return v1
        elif tamanho == 6:
            v1 = f'{valor[0]}.{valor[1]}{valor[2]}{valor[3]},{valor[4]}{valor[5]}'
            print(v1)
            return v1
        elif tamanho == 5:
            v1 = f'{valor[0]}{valor[1]}{valor[2]},{valor[3]}{valor[4]}'
            print(v1)
            return v1
        elif tamanho == 4:
            v1 = f'{valor[0]}{valor[1]},{valor[2]}{valor[3]}'
            print(v1)
            return v1
        else:
            print('erro')
    else:
        filtro1 = varr.replace(',', '').replace('.', '')
        tamanho = len(str(filtro1))
        print(tamanho)

        valor = str(filtro1)
        # print("valor", valor)
        if tamanho == 9:
            v1 = f'{valor[0]}.{valor[1]}{valor[2]}{valor[3]}.{valor[4]}{valor[5]}{valor[6]},{valor[7]}{valor[8]}'
            print(v1)
            return v1
        elif tamanho == 8:
            v1 = f'{valor[0]}{valor[1]}{valor[2]}.{valor[3]}{valor[4]}{valor[5]},{valor[6]}{valor[7]}'
            print(v1)
            return v1
        elif tamanho == 7:
            v1 = f'{valor[0]}{valor[1]}.{valor[2]}{valor[3]}{valor[4]},{valor[5]}{valor[6]}'
            print(v1)
            return v1
        elif tamanho == 6:
            v1 = f'{valor[0]}.{valor[1]}{valor[2]}{valor[3]},{valor[4]}{valor[5]}'
            print(v1)
            return v1
        elif tamanho == 5:
            v1 = f'{valor[0]}{valor[1]}{valor[2]},{valor[3]}{valor[4]}'
            print(v1)
            return v1
        elif tamanho == 4:
            v1 = f'{valor[0]}{valor[1]},{valor[2]}{valor[3]}'
            print(v1)
            return v1
        elif tamanho == 3:
            v1 = f'0{valor[0]},{valor[1]}{valor[2]}'
            print(v1)
            return v1
        elif tamanho == 2:
            v1 = f'00,{valor[0]}{valor[1]}'
            print(v1)
            return v1
        elif tamanho == 1:
            v1 = f'00,0{valor[0]}'
            print(v1)
            return v1
        else:
            print('erro')


def entry_corrigir(varr):
    varr = str(varr)
    if ',' in varr or '.' in varr:
        filtro1 = varr.replace(',', '').replace('.', '')
        tamanho = len(str(filtro1))
        valor = str(filtro1)
       # print("valor", valor)
        if tamanho == 9:
            v1 = f'{valor[0]}.{valor[1]}{valor[2]}{valor[3]}.{valor[4]}{valor[5]}{valor[6]},{valor[7]}{valor[8]}'
            print(v1)
            return v1
        elif tamanho == 8:
            v1 = f'{valor[0]}{valor[1]}{valor[2]}.{valor[3]}{valor[4]}{valor[5]},{valor[6]}{valor[7]}'
            print(v1)
            return v1
        elif tamanho == 7:
            v1 = f'{valor[0]}{valor[1]}.{valor[2]}{valor[3]}{valor[4]},{valor[5]}{valor[6]}'
            print(v1)
            return v1
        elif tamanho == 6:
            v1 = f'{valor[0]}.{valor[1]}{valor[2]}{valor[3]},{valor[4]}{valor[5]}'
            print(v1)
            return v1
        elif tamanho == 5:
            v1 = f'{valor[0]}{valor[1]}{valor[2]},{valor[3]}{valor[4]}'
            print(v1)
            return v1
        elif tamanho == 4:
            v1 = f'{valor[0]}{valor[1]},{valor[2]}{valor[3]}'
            print(v1)
            return v1
        elif tamanho == 3:
            v1 = f'0{valor[0]},{valor[1]}{valor[2]}'
            print(v1)
            return v1

        else:
            print('erro')
    else:
        filtro1 = varr.replace(',', '').replace('.', '')
        tamanho = len(str(filtro1))
        valor = str(filtro1)
        print("valor", valor)

        if tamanho == 9:
            # 100.000,000,00
            v1 = f'{valor[0]}{valor[1]}{valor[2]}.{valor[3]}{valor[4]}{valor[5]}.{valor[6]}{valor[7]}{valor[8]},00'
            print(v1)
            return v1
        elif tamanho == 8:
            # 10.000.000,00
            v1 = f'{valor[0]}{valor[1]}.{valor[2]}{valor[3]}{valor[4]}.{valor[5]}{valor[6]}{valor[7]},00'
            print(v1)
            return v1
        elif tamanho == 7:
            # 1.000.000,00
            v1 = f'{valor[0]}.{valor[1]}{valor[2]}{valor[3]}.{valor[4]}{valor[5]}{valor[6]},00'
            print(v1)
            return v1
        elif tamanho == 6:
            # 100.000,00
            v1 = f'{valor[0]}{valor[1]}{valor[2]}.{valor[3]}{valor[4]}{valor[5]},00'
            print(v1)
            return v1
        elif tamanho == 5:
            # 10.000,00
            v1 = f'{valor[0]}{valor[1]}.{valor[2]}{valor[3]}{valor[4]},00'
            print(v1)
            return v1
        elif tamanho == 4:
            v1 = f'{valor[0]}.{valor[1]}{valor[2]}{valor[3]},00'  # 1.000,00
            print(v1)
            return v1
        elif tamanho == 3:
            v1 = f'{valor[0]}{valor[1]}{valor[2]},00'  # 100,00
            print(v1)
            return v1
        elif tamanho == 2:
            v1 = f'{valor[0]}{valor[1]},00'  # 10,00
            print(v1)
            return v1
        elif tamanho == 1:
            v1 = f'0{valor[0]},00'  # 01,00
            print(v1)
            return v1
        else:
            print('erro')
###################################################################################################################
#                                                           funcões dos botões


def tree_att():  # atualizar treeview principal
    tree.delete(*tree.get_children())
    conn = sqlite3.connect('elementos/doces.db')
    c = conn.cursor()
    c.execute("SELECT * FROM teste")
    testes = c.fetchall()
    global count
    count = 0
    for record in testes:
        if count % 2 == 0:
            tree.insert(parent='', index='end', iid=count, text='', values=(
                record[0], record[1], record[2], record[3], record[4], 'R$'+str(record[5])), tags=('evenrow',))
        else:
            pass
            tree.insert(parent='', index='end', iid=count, text='', values=(
                record[0], record[1], record[2], record[3], record[4], 'R$'+str(record[5])), tags=('oddrow',))
        count += 1
    conn.commit()
    conn.close()


def salvarr():  # save
    if ent1.get() == '':
        lblinfos('escreva um nome para salvar!', 'red')
        return
    elif ent2s.get() == '':
        lblinfos('escreva um produto para salvar!', 'red')
        return
    elif ent3.get() == '':
        lblinfos('escreva uma quantidade para salvar!', 'red')
        return
    else:

        ress = []

        def produtos_ligado(produti):
            ress.clear()
            conn = sqlite3.connect('elementos/doces.db')
            c = conn.cursor()
            c.execute(f"SELECT * FROM produtos")
            produtos = c.fetchall()
            # print(produtos)
            # print('''
            # lista de todos produtos''')
            product = []
            for produtoz in produtos:
                product.append(produtoz[0])

            preços = []
            for price in produtos:
                preços.append(price[1])

            # print(preços)
            # print('''
                # lista de todos preços''')

            # print(product)
            # print('''
                # lista de todos nomes de produtos''')
            dici = dict(zip(product, preços))

            # print(dici)
            # print('''
            # lista de todos produtos em dicionario''')
            # print(dici[produti])

            ress.append(dici[produti])

            conn.commit()
            conn.close()
            # print(f'o valor do resultado é{ress[0]}')

            # print(f'o valor resultado era {ress}')
            # print('''
            # o valor da lista atual agora é''')

        produtos_ligado(ent2s.get())

        nmr1 = ress[0].replace(',', '')
        nmr2 = ent3.get()
        vezz = int(nmr1)*int(nmr2)
        print(nmr1)
        print(nmr2)
        print(vezz)

        print(corrigir_resultado(vezz))

        if str.isnumeric(ent3.get()):
            # tree.delete(*tree.get_children())
            conn = sqlite3.connect('elementos/doces.db')
            cursor = conn.cursor()
            cursor.execute("INSERT INTO teste (nome, produto, quantidade, data, preço) VALUES (?,?,?,?,?)",
                           (ent1.get(), ent2s.get(), ent3.get(), hoje, corrigir_resultado(vezz)))
            # cursor.execute('UPDATE produtos SET quantidade = quantidade - ? WHERE produtos = ?', (ent2s.get(), ent3.get(),))
            cursor.execute('UPDATE produtos SET quantidade = quantidade - ? WHERE produtos = ?', (ent3.get(), ent2s.get()))
            conn.commit()
            conn.close()
            ent1.delete(0, 'end')
            ent2s.delete(0, 'end')
            ent3.delete(0, 'end')
            lblinfos('item salvo!', verde)
            tree_att()
        else:
            lblinfos('escreva uma quantidade numerica!', 'red')


def salvarrr():
    salvarr()





def informaçoess():  # label com informaçoes iniciais
    try:
        items1 = tree.selection()[0]
        v = tree.item(items1, 'values')

        lblinfos(
            f"no dia {data_verse} {v[0]} pegou {v[2]} {v[1]}, que ficou \n no valor de {v[5]} reais. ", verde)
    except Exception as error:
        lblinfos('selecione um item para obter informações!', 'red')
        print(error)


def anotaçoesz():  # abrir anotaçoes
    os.startfile(
        'C:/Users/jackson/Desktop/aplicativos em trabalho/ericaapp/elementos/erica_anot.txt')


###################################################################################################################
#                                                         def botões


def btapp(btapp, foto, xp, yp, comando, cor):  # botoes premoldados
    btapp = Button(tela, image=foto, command=comando)
    btapp.configure(bd=0, background=cor, activebackground=cor)
    btapp.place(x=xp, y=yp)
###################################################################################################################


def btfacebook():  # face
    wb.open('https://www.facebook.com/erica.gomes.942145')


def btwpp():  # abrir wpp web
    wb.open('https://web.whatsapp.com')
   

def eudora():  # site eudora
    wb.open('https://www.eudora.com.br/?gclid=CjwKCAjwv8qkBhAnEiwAkY-ahtOb2c6Cby1X2sO7eTRSdcoFSmFTEj337Y7Gxu6QSj3SSsqPClZeJxoCTYsQAvD_BwE&gclsrc=aw.ds')


def btinstagram():  # insta
    wb.open('https://www.instagram.com/eudora.134/?igshid=MzNlNGNkZWQ4Mg%3D%3D')


def btparcialmente():  # pagar parcialmente uma conta

    def deletarr():  # delete

        def salvar_excluido():
            # abrir arquivo do Excel
            wb = openpyxl.load_workbook('elementos\historico_clientes.xlsx')
            # selecionar planilha do Excel
            ws = wb.active

            # criar nova linha na planilha
            linha = ws.max_row + 1

            # inserir valores nas colunas relevantes
            ws.cell(row=linha, column=1).value = data_verse
            ws.cell(row=linha, column=2).value = b[0]
            ws.cell(row=linha, column=3).value = b[1]
            ws.cell(row=linha, column=4).value = b[2]
            ws.cell(row=linha, column=5).value = b[5]
            # salvar arquivo do Excel
            wb.save('elementos\historico_clientes.xlsx')

        try:
            itemsel = treez.selection()[0]
            itemselx = treez.selection()
            b = treez.item(itemsel, 'values')
            bbc = tree.item(itemselx, 'values')
            print(bbc[1])

            salvar_excluido()
            # print(f'{b[0]},{b[1]},{b[2]}')
            numeric = int(itemsel)
            conn = sqlite3.connect('elementos/doces.db')
            c = conn.cursor()
            c.execute("SELECT id FROM teste")
            testes = c.fetchall()
            conn = sqlite3.connect('elementos/doces.db')
            c = conn.cursor()
            c.execute("delete from teste where id = ?", (testes[numeric]))
            conn.commit()
            conn.close()

            treez.delete(*treez.get_children())
            tree_att()
            ss()
            lblinfos('item deletado!', verde)


        except:
            print(Exception)
            lblinfos('escolha um item para deletar', 'red')   


    def editar():
        selection = treez.selection()
        if selection:
            values = treez.item(selection)['values']

            valor1 = values[1]  # nome produto

            valor2 = values[2]  # devendo

            valor3 = values[3]  # pago

            valor5 = values[5]  # id

        else:
            print("Nenhum item selecionado")

        resultado_entry = entry_corrigir(mais_ent1.get())
        resultado_entry_filtro = resultado_entry.replace(
            ',', '').replace('.', '').replace('R$', '')

        valor_entry = int(resultado_entry_filtro)

        resultado_devido = valor2.replace(
            ',', '').replace('.', '').replace('R$', '')
        valor_devido = int(resultado_devido)

        resultado_pago = valor3.replace(
            ',', '').replace('.', '').replace('R$', '')
        valor_pago = int(resultado_pago)

        print(
            f'o valor pago é {valor_pago} e o valor que ainda deve é {valor_devido}')

        novo_valor_devido = int(valor_devido) - int(valor_entry)

        novo_valor_pago = int(valor_pago) + int(valor_entry)

        print(
            f'o valor novo pago é {novo_valor_pago} e o valor que ainda deve é {novo_valor_devido}')

        a = corrigir_resultado(novo_valor_devido)
        print(f'error a {a}')
        # aqui esta o erro !!!!!!!!!!!!!!!!!!!11111111
        b = corrigir_resultado(novo_valor_pago)
        print(f'error b {b}')
        cc = valor5
        print(f'procurando o erro com o a {a}')
        print(
            f'-----------------------------------------------------{novo_valor_pago}')
        # /estudando o erro (b)

        print(f'o valor com erro é o valor {b}')

        if novo_valor_devido <= 0:
            print('zerado')
            conn = sqlite3.connect('elementos/doces.db')
            c = conn.cursor()
            c.execute(
                'UPDATE produtos SET quantidade = quantidade - 1 WHERE produtos = ?', (valor1,))
            c.execute("DELETE FROM teste WHERE id = ?", (cc,))
            conn.commit()
            conn.close()
            treez.delete(*treez.get_children())
            ss()
        else:
            print('com valor')
            conn = sqlite3.connect('elementos/doces.db')
            c = conn.cursor()
            # o erro esta por aqui/ no (b)
            print(
                f'o novo valor devido é {a} e o novo pago é {b} e o id é {cc}')
            sql = "UPDATE teste SET preço = ?, `pago parcial` = ? WHERE id = ?"
            c.execute(sql, (a, b, cc))
            conn.commit()
            conn.close()
            treez.delete(*treez.get_children())
            ss()
        tree_att()

    def ss():
        treez.delete(*treez.get_children())
        conn = sqlite3.connect('elementos/doces.db')
        c = conn.cursor()
        c.execute("SELECT * FROM teste")
        testesz = c.fetchall()
        global count
        count = 0
        for record in testesz:
            if count % 2 == 0:
                treez.insert(parent='', index='end', iid=count, text='', values=(
                    record[0], record[1], record[5], record[6], record[3], record[4]), tags=('evenrow',))
            else:
                pass
                treez.insert(parent='', index='end', iid=count, text='', values=(
                    record[0], record[1], record[5], record[6], record[3], record[4]), tags=('oddrow',))
            count += 1
        conn.commit()
        conn.close()

    ab = Toplevel(tela1)
    ab.resizable(False, False)
    ab.iconbitmap('front_end/botoes_redes/E.ico')
    ab.title('pago parcialmente')

    ab.configure(bg='#BBE0E5')
    ab.geometry('600x545')
    treef = Frame(ab)
    treef.pack()
    treef.configure(bg='lightgrey')
    tree_scroll = Scrollbar(treef)
    tree_scroll.pack(side=RIGHT, fill=Y)
    treez = ttk.Treeview(treef, show='headings',
                         yscrollcommand=tree_scroll.set, selectmode="extended")
    treez.configure(height=6)
    treez.pack()
    s = ttk.Style()
    s.configure('Treeview', rowheight=45, font=('arial', 15))
    treez['columns'] = ('nome', 'produto', 'preço', 'pago', 'data', 'id')
    treez.heading("#0", text="", anchor=W)
    treez.heading('nome', text='nome')
    treez.heading('produto', text='produto')
    treez.heading('preço', text='devendo')
    treez.heading('pago', text='pago')
    treez.heading('data', text='data')
    treez.heading('id', text='id')
    treez.column('nome', width=145, anchor='c')
    treez.column('produto', width=100, anchor='c')
    treez.column('preço', width=100, anchor='c')
    treez.column('pago', width=100, anchor='c')
    treez.column('data', width=100, anchor='c')
    treez.column('id', width=32, anchor='c')

    tree_scroll.config(command=treez.yview)
    treez.tag_configure('oddrow', background='#4386C8', foreground='lightgrey')
    treez.tag_configure('evenrow', background='lightgrey',
                        foreground='#4386C8')
    ss()
    tree.delete()

    abc = Frame(ab)
    abc.configure(bg='#BBE0E5')
    abc.pack()

    mais_label1 = Label(abc, image=valor_pago1, bg='#009688')
    mais_label1.pack()

    mais_ent1 = Entry(abc, font=('(arial', 27))
    mais_ent1.configure(width=30)
    mais_ent1.pack()

    mais_btt = Button(abc, image=salvarparcialmente,
                      command=editar, bg='#4D708B')
    mais_btt.pack()

    mns_btt = Button(abc, image=pagoutodo,
                      command=deletarr, bg='#4D708B')
    mns_btt.pack()
    conex()


def btconfig():  # Adicionar ou excluir produtos
    def bt_dltt():
        try:
            itemsel = tree.selection()[0]
            numeric = int(itemsel)
            conn = sqlite3.connect('elementos/doces.db')
            c = conn.cursor()
            c.execute("SELECT produtos FROM produtos")
            testes = c.fetchall()
            conn = sqlite3.connect('elementos/doces.db')
            c = conn.cursor()
            c.execute("delete from produtos where produtos = ?",
                      (testes[numeric]))
            conn.commit()
            conn.close()

            print('produto deletado')
            ss()
            conex()

        except:
            print('IMPOSIVEL DELETAR')

    def ss():
        tree.delete(*tree.get_children())
        conn = sqlite3.connect('elementos/doces.db')
        c = conn.cursor()
        c.execute("SELECT * FROM produtos")
        testesz = c.fetchall()
        global count
        count = 0
        for record in testesz:
            if count % 2 == 0:
                tree.insert(parent='', index='end', iid=count, text='', values=(
                    record[0], 'R$'+str(record[1]), record[2]), tags=('evenrow',))
            else:
                pass
                tree.insert(parent='', index='end', iid=count, text='', values=(
                    record[0], 'R$'+str(record[1]), record[2]), tags=('oddrow',))
            count += 1
        conn.commit()
        conn.close()

    ab = Toplevel(tela1)
    ab.resizable(False, False)
    ab.iconbitmap('front_end/botoes_redes/E.ico')
    ab.title('adicionar produtos')
    ab.configure(bg='#BBE0E5')
    ab.geometry('420x610')
    treef = Frame(ab)
    treef.pack()
    treef.configure(bg='lightgrey')
    tree_scroll = Scrollbar(treef)
    tree_scroll.pack(side=RIGHT, fill=Y)
    tree = ttk.Treeview(treef, show='headings',
                        yscrollcommand=tree_scroll.set, selectmode="extended")
    tree.configure(height=6)
    tree.pack()
    s = ttk.Style()
    s.configure('Treeview', rowheight=45, font=('arial', 15))
    tree['columns'] = ('produto', 'preço')
    tree.heading("#0", text="", anchor=W)
    tree.heading('produto', text='produto')
    tree.heading('preço', text='preço')

    tree.column('produto', width=290, anchor='c')
    tree.column('preço', width=100, anchor='c')

    tree_scroll.config(command=tree.yview)
    tree.tag_configure('oddrow', background='#4386C8', foreground='lightgrey')
    tree.tag_configure('evenrow', background='lightgrey', foreground='#4386C8')
    ss()
#########################################################################

    def add():
        if mais_ent1.get() == '':
            print('error')
            return
        elif mais_ent2.get() == '':
            print('error')
            return
        else:
            avea = entry_corrigir(mais_ent2.get())
            produtos.adicionar_produtos(mais_ent1.get(), avea)
            mais_ent1.delete(0, 'end')
            mais_ent2.delete(0, 'end')
            tree.delete(*tree.get_children())
            ss()
            conex()

    abc = Frame(ab)
    abc.configure(bg='#BBE0E5')
    abc.pack()

    mais_label1 = Label(abc, image=produto, bg='#BBE0E5')
    mais_label1.pack()

    mais_ent1 = Entry(abc, font=('(arial', 27))
    mais_ent1.pack()

    mais_label2 = Label(abc, image=preço, bg='#BBE0E5')
    mais_label2.pack()

    mais_ent2 = Entry(abc, font=('(arial', 27))
    mais_ent2.pack()

    mais_btt = Button(abc, image=salvar, command=add,
                      bg='#BBE0E5', bd=0, pady=50)
    mais_btt.pack()

    menos_btt = Button(abc, image=deletarrx, command=bt_dltt,
                       bg='#BBE0E5', bd=0, pady=50)
    menos_btt.pack()
    conex()


def bthistorico():  # Excel
    os.startfile(
        'elementos\historico_clientes.xlsx')


def devidendos():  # valor devidos

    def pesquisa():
        treezc.delete(*treezc.get_children())
        conn = sqlite3.connect('elementos/doces.db')
        c = conn.cursor()
        c.execute("SELECT * FROM teste WHERE nome = ?", (mais_ent1s.get(),))
        testeszc = c.fetchall()
        global count
        count = 0
        for record in testeszc:
            if count % 2 == 0:
                treezc.insert(parent='', index='end', iid=count, text='', values=(
                    record[0], 'R$ '+str(record[5]), 'R$ '+str(record[6]), record[3]), tags=('evenrow',))
            else:
                pass
                treezc.insert(parent='', index='end', iid=count, text='', values=(
                    record[0], 'R$ '+str(record[5]), 'R$ '+str(record[6]), record[3]), tags=('oddrow',))
            count += 1
        conn.commit()
        conn.close()
        calcular_soma()
        mais_ent1s.delete(0, 'end')

    def calcular_soma():
        total = 0
        for item in treezc.get_children():
            valor = str(treezc.item(item, 'values')[1])
            a = valor.replace(',', '').replace('.', '').replace('R$', '')
            print(a)

            total += int(a)

        print(total)

        total = str(total)
        ros = corrigir_resultado(total)
        print(ros)
        lab["text"] = f"Soma de valores devidos: R$ {ros}."

    def v():
        ss()
        lab["text"] = f"escreva um nome para informações."

    def ss():
        treezc.delete(*treezc.get_children())
        conn = sqlite3.connect('elementos/doces.db')
        c = conn.cursor()
        c.execute("SELECT * FROM teste")
        testeszc = c.fetchall()
        global count
        count = 0
        for record in testeszc:
            if count % 2 == 0:
                treezc.insert(parent='', index='end', iid=count, text='', values=(
                    record[0],  'R$ '+str(record[5]),  'R$ '+str(record[6]), record[3]), tags=('evenrow',))
            else:
                pass
                treezc.insert(parent='', index='end', iid=count, text='', values=(
                    record[0],  'R$ '+record[5],  'R$ '+str(record[6]), record[3]), tags=('oddrow',))
            count += 1
        conn.commit()
        conn.close()

    ab = Toplevel(tela1)
    ab.resizable(False, False)
    ab.iconbitmap('front_end/botoes_redes/E.ico')
    ab.title('valor a pagar')

    ab.configure(bg='#BBE0E5')
    ab.geometry('600x570')
    treef = Frame(ab)
    treef.pack()
    treef.configure(bg='lightgrey')
    tree_scroll = Scrollbar(treef)
    tree_scroll.pack(side=RIGHT, fill=Y)
    treezc = ttk.Treeview(treef, show='headings',
                          yscrollcommand=tree_scroll.set, selectmode="extended")
    treezc.configure(height=6)
    treezc.pack()
    s = ttk.Style()
    s.configure('Treeview', rowheight=45, font=('arial', 15))
    treezc['columns'] = ('nome', 'preço', 'pago', 'data')
    treezc.heading("#0", text="", anchor=W)
    treezc.heading('nome', text='nome')
    treezc.heading('preço', text='devendo')
    treezc.heading('pago', text='pago')
    treezc.heading('data', text='data')
    treezc.column('nome', width=150, anchor='c')
    treezc.column('preço', width=100, anchor='c')
    treezc.column('pago', width=100, anchor='c')
    treezc.column('data', width=200, anchor='c')

    tree_scroll.config(command=treezc.yview)
    treezc.tag_configure('oddrow', background='#4386C8',
                         foreground='lightgrey')
    treezc.tag_configure('evenrow', background='lightgrey',
                         foreground='#4386C8')
    ss()

    abc = Frame(ab)
    abc.configure(bg='#009688')
    abc.pack()

    mais_label1 = Label(abc, image=valor_pago, bg='#009688')
    mais_label1.pack()

    lab = Label(abc, text='escreva um nome para infonformações.')
    lab.configure(bg='#009688', fg='white', font=('arial, 20'), width=35)
    lab.pack()

    mais_ent1s = Entry(abc, font=('(arial', 27))
    mais_ent1s.configure(width=28)
    mais_ent1s.pack()

    mais_btts = Button(abc, image=pesquisar,
                       command=pesquisa, bg='#4D708B')
    mais_btts.pack()

    mais_bttsx = Button(abc, image=limpar_pesquisa,
                        command=v, bg='#4D708B')
    mais_bttsx.pack()

    conex()


def estoque():
    def alterar():
        selection = treezc.selection()
        values = treezc.item(selection)['values']
        idd = values[3]
        conn = sqlite3.connect('elementos/doces.db')
        c = conn.cursor()
        c.execute('UPDATE produtos SET quantidade = ? WHERE id = ?',
                  (mais_ent1s.get(), idd))
        conn.commit()
        conn.close()
        treezc.delete(*treezc.get_children())
        ss()
        mais_ent1s.delete(0, 'end')

    def ss():
        treezc.delete(*treezc.get_children())
        conn = sqlite3.connect('elementos/doces.db')
        c = conn.cursor()
        c.execute("SELECT * FROM produtos")
        testeszc = c.fetchall()
        global count
        count = 0
        for record in testeszc:
            if count % 2 == 0:
                treezc.insert(parent='', index='end', iid=count, text='', values=(
                    record[0],  (record[1]),  (record[2]), record[3]), tags=('evenrow',))
            else:
                pass
                treezc.insert(parent='', index='end', iid=count, text='', values=(
                    record[0],  record[1],  record[2], record[3]), tags=('oddrow',))
            count += 1
        conn.commit()
        conn.close()

    ab = Toplevel(tela1)
    ab.resizable(False, False)
    ab.iconbitmap('front_end/botoes_redes/E.ico')
    ab.title('ajuste e consulta de estoque')

    ab.configure(bg='#BBE0E5')
    ab.geometry('600x455')
    treef = Frame(ab)
    treef.pack()
    treef.configure(bg='lightgrey')
    tree_scroll = Scrollbar(treef)
    tree_scroll.pack(side=RIGHT, fill=Y)
    treezc = ttk.Treeview(treef, show='headings',
                          yscrollcommand=tree_scroll.set, selectmode="extended")
    treezc.configure(height=6)
    treezc.pack()
    s = ttk.Style()
    s.configure('Treeview', rowheight=45, font=('arial', 15))
    treezc['columns'] = ('produto', 'preço', 'qtd', 'id')
    treezc.heading("#0", text="", anchor=W)
    treezc.heading('produto', text='produto')
    treezc.heading('preço', text='preço')
    treezc.heading('qtd', text='qtd')
    treezc.heading('id', text='id')
    treezc.column('produto', width=180, anchor='c')
    treezc.column('preço', width=100, anchor='c')
    treezc.column('qtd', width=100, anchor='c')
    treezc.column('id', width=200, anchor='c')

    tree_scroll.config(command=treezc.yview)
    treezc.tag_configure('oddrow', background='#4386C8',
                         foreground='lightgrey')
    treezc.tag_configure('evenrow', background='lightgrey',
                         foreground='#4386C8')
    ss()

    abc = Frame(ab)
    abc.configure(bg='#009688')
    abc.pack()

    mais_label1 = Label(abc, image=atututu, bg='#009688')
    mais_label1.pack()

    mais_ent1s = Entry(abc, font=('(arial', 27))
    mais_ent1s.configure(width=28)
    mais_ent1s.pack()

    mais_btts = Button(abc, image=atutu,
                       command=alterar, bg='#4D708B')
    mais_btts.pack()
###################################################################################################################
#                                                           tela principal
tela1.title('Erica App')
tela1.geometry('1366x720+0+0')
tela1.iconbitmap('front_end/botoes_redes/E.ico')
###################################################################################################################
###################################################################################################################
#                                                           frame principal
tela = Frame(tela1)
tela.place(x=0, y=0)
fundoft = PhotoImage(file='front_end/fundo/projeto.png')
fundo = Label(tela, image=fundoft)
fundo.pack()
###################################################################################################################
###################################################################################################################
#                                                           funçoes imagem e label


def lblinfos(oque, cor):
    info = Label(tela, text=oque, font=('arial', 18), bg=cor, fg='white')
    info.configure(width=59, height=3)
    info.place(x=265, y=281)


lblinfos('escolha algum item do treeview! ', '#009688')
###################################################################################################################
###################################################################################################################
#                                                           entrys
ent1 = Entry(tela, font=('(arial', 29))
ent1.configure(width=12)
ent1.place(x=268, y=92)
conex()
ent3 = Entry(tela, font=('(arial', 29))
ent3.configure(width=12)
ent3.place(x=825, y=92)
###################################################################################################################
###################################################################################################################
#                                                           botões na tela
btapp('salvar', salvar, 265, 150, salvarrr, '#BBE0E5')
btapp('deletar', deletar, 685, 150, mudar_tema, '#BBE0E5')
btapp('informaçoes', informaçoes, 265, 210, informaçoess, '#BBE0E5')
btapp('anotaçoes', anotaçoes, 685, 210, anotaçoesz, '#BBE0E5')
#                                                           botões na tela rede
btapp('facebook', facebook, 1145, 30, btfacebook, '#4386C8')
btapp('whatsapp', whatsapp, 1220, 30, btwpp, '#4386C8')
btapp('instagram', instagram, 1287, 30, btinstagram, '#4386C8')
btapp('moeda', moeda, 1145, 150, btparcialmente, '#4386C8')
btapp('config', atualizar_preços, 1220, 150, btconfig, '#4386C8')
btapp('bthistorico', historico, 1289, 95, bthistorico, '#4386C8')
btapp('perfume', perfume, 1145, 95, estoque, '#4386C8')
btapp('e', e, 1220, 95, eudora, '#4386C8')
btapp('bd', bd, 1287, 150, devidendos, '#4386C8')

###################################################################################################################


foto1 = Label(tela, bg='#BBE0E5')
foto2 = Label(tela, bg='#BBE0E5')
foto3 = Label(tela, bg='#BBE0E5')
foto4 = Label(tela, bg='#BBE0E5')

############################


foto1.place(x=15, y=15)
foto2.place(x=15, y=242)
foto3.place(x=15, y=475)
foto4.place(x=1122, y=237)


###################################################################################################################
#                                                           treeview
treef = Frame(tela)
treef.place(x=273, y=403)
treef.configure(bg='pink')
tree_scroll = Scrollbar(treef)
tree_scroll.pack(side=RIGHT, fill=Y)
tree = ttk.Treeview(treef, show='headings',
                    yscrollcommand=tree_scroll.set, selectmode="extended")
tree.configure(height=6)
tree.pack()
s = ttk.Style()
s.configure('Treeview', rowheight=45, font=('arial', 15))
tree['columns'] = ('nome', 'produto', 'quantidade', 'data', 'id', 'preço')
tree.heading("#0", text="", anchor=W)
tree.heading('nome', text='nome')
tree.heading('produto', text='produto')
tree.heading('quantidade', text='qtd')
tree.heading('preço', text='total')
tree.heading('data', text='data')
tree.heading('id', text='pedido')
tree.column('nome', width=226, anchor='c')
tree.column('produto', width=215, anchor='c')
tree.column('quantidade', width=60, anchor='c')
tree.column('preço', width=110, anchor='c')
tree.column('data', width=130, anchor='c')
tree.column('id', width=60, anchor='c')
tree_scroll.config(command=tree.yview)
tree.tag_configure('oddrow', background='#4386C8', foreground='lightgrey')
tree.tag_configure('evenrow', background='lightgrey', foreground='#4386C8')
###################################################################################################################
#                                                           run app
load()
tree_att()
tela1.mainloop()
